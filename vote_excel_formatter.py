from __future__ import annotations

import argparse
import copy
import io
import re
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from pypdf import PdfReader
from pypdf.generic import ContentStream
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
GROUP_FILL = PatternFill("solid", fgColor="D9EAF7")
SIZE_FILL = PatternFill("solid", fgColor="EAF2F8")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass
class ProductRow:
    row_index: int
    seq: str
    order_no: str
    customer: str
    style_no: str
    name: str
    color: str
    quantities: list[float | int | str | None]
    total: float | int | str | None


@dataclass
class ProductGroup:
    style_no: str
    name: str
    rows: list[ProductRow] = field(default_factory=list)
    image_bytes: bytes | None = None


SIZE_HEADERS = [
    "KG / 均码 / 25 / 36",
    "S / 26 / 37",
    "M / 27 / 38",
    "L / 28 / 39",
    "4XL / 29 / 40",
    "2XL / 30 / 41",
    "XL / 31 / 35",
    "3XL / 32 / 34",
    "1500",
    "150-180",
]
SIZE_X_POSITIONS = [340.0, 383.6, 427.2, 470.7, 514.3, 557.9, 601.5, 645.0, 688.6, 732.2]
DATA_ROW_HEIGHT_PT = 28.35
DATA_ROW_HEIGHT_PX = 38


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def parse_number(value: Any) -> int | float | str | None:
    text = clean_text(value)
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return value
    if number.is_integer():
        return int(number)
    return number


def numeric_total(values: list[Any]) -> float:
    total = 0.0
    for value in values:
        if isinstance(value, (int, float)):
            total += value
        elif isinstance(value, str):
            try:
                total += float(value)
            except ValueError:
                pass
    return total


def hide_zero(value: Any) -> Any:
    if value == 0 or value == 0.0 or clean_text(value) == "0":
        return None
    return value


def find_header_columns(ws) -> dict[str, int]:
    header = {clean_text(cell.value): cell.column for cell in ws[1] if clean_text(cell.value)}
    required = ["图片", "款号", "名称", "颜色", "订货数"]
    missing = [name for name in required if name not in header]
    if missing:
        raise ValueError(f"找不到必要表头：{', '.join(missing)}")
    return header


def read_size_headers(ws, start_col: int, end_col: int) -> list[str]:
    headers: list[str] = []
    for col in range(start_col, end_col + 1):
        parts = []
        for row in range(1, 5):
            text = clean_text(ws.cell(row=row, column=col).value)
            if text:
                parts.append(text)
        label = " / ".join(parts)
        headers.append(label or f"尺码{col - start_col + 1}")
    return headers


def extract_images_by_row(ws) -> dict[int, bytes]:
    images: dict[int, bytes] = {}
    for image in getattr(ws, "_images", []):
        try:
            row = image.anchor._from.row + 1
            images.setdefault(row, image._data())
        except Exception:
            continue
    return images


def load_groups(input_path: Path) -> tuple[list[str], OrderedDict[str, ProductGroup]]:
    workbook = openpyxl.load_workbook(input_path)
    ws = workbook.active

    columns = find_header_columns(ws)
    order_col = columns.get("订单号")
    customer_col = columns.get("客户")
    style_col = columns["款号"]
    name_col = columns["名称"]
    color_col = columns["颜色"]
    total_col = columns["订货数"]

    first_size_col = color_col + 1
    last_size_col = total_col - 1
    size_headers = read_size_headers(ws, first_size_col, last_size_col)
    images_by_row = extract_images_by_row(ws)

    groups: OrderedDict[str, ProductGroup] = OrderedDict()
    for row_index in range(5, ws.max_row + 1):
        style_no = clean_text(ws.cell(row_index, style_col).value)
        order_no = clean_text(ws.cell(row_index, order_col).value) if order_col else ""
        customer = clean_text(ws.cell(row_index, customer_col).value) if customer_col else ""
        name = clean_text(ws.cell(row_index, name_col).value)
        color = clean_text(ws.cell(row_index, color_col).value)
        if not style_no or style_no == "00000" or name == "抹零":
            continue

        quantities = [ws.cell(row_index, col).value for col in range(first_size_col, last_size_col + 1)]
        total = ws.cell(row_index, total_col).value
        if numeric_total(quantities) <= 0 and not total:
            continue

        group = groups.setdefault(style_no, ProductGroup(style_no=style_no, name=name))
        if not group.name and name:
            group.name = name
        if group.image_bytes is None:
            group.image_bytes = images_by_row.get(row_index)
        group.rows.append(
            ProductRow(
                row_index=row_index,
                seq="",
                order_no=order_no,
                customer=customer,
                style_no=style_no,
                name=name,
                color=color,
                quantities=quantities,
                total=total,
            )
        )

    # If the first row for a款号 had no image, look for any later image within the group.
    for group in groups.values():
        if group.image_bytes is None:
            for product_row in group.rows:
                if product_row.row_index in images_by_row:
                    group.image_bytes = images_by_row[product_row.row_index]
                    break

    return size_headers, groups


def load_vertical_xlsx_rows(input_path: Path) -> tuple[list[str], list[ProductRow], dict[str, bytes]]:
    workbook = openpyxl.load_workbook(input_path)
    ws = workbook.active
    columns = {clean_text(cell.value): cell.column for cell in ws[1] if clean_text(cell.value)}
    required = ["批次", "客户", "图片", "款号", "名称", "颜色", "尺码", "订货数"]
    missing = [name for name in required if name not in columns]
    if missing:
        raise ValueError(f"找不到必要表头：{', '.join(missing)}")

    images_by_row = extract_images_by_row(ws)
    sizes: list[str] = []
    grouped: OrderedDict[tuple[str, str, str, str, str, str], ProductRow] = OrderedDict()
    images_by_style: dict[str, bytes] = {}

    for row_index in range(2, ws.max_row + 1):
        style_no = clean_text(ws.cell(row_index, columns["款号"]).value)
        name = clean_text(ws.cell(row_index, columns["名称"]).value)
        color = clean_text(ws.cell(row_index, columns["颜色"]).value)
        size = clean_text(ws.cell(row_index, columns["尺码"]).value)
        quantity = parse_number(ws.cell(row_index, columns["订货数"]).value)
        order_no = clean_text(ws.cell(row_index, columns["批次"]).value)
        customer = clean_text(ws.cell(row_index, columns["客户"]).value)
        if not style_no or not size or quantity in (None, ""):
            continue
        if style_no == "00000" or name == "抹零":
            continue
        if size not in sizes:
            sizes.append(size)

        key = (order_no, customer, style_no, name, color, "")
        product_row = grouped.get(key)
        if product_row is None:
            product_row = ProductRow(
                row_index=row_index,
                seq=str(len(grouped) + 1),
                order_no=order_no,
                customer=customer,
                style_no=style_no,
                name=name,
                color=color,
                quantities=[],
                total=0,
            )
            grouped[key] = product_row

        while len(product_row.quantities) < len(sizes):
            product_row.quantities.append(None)
        size_index = sizes.index(size)
        existing = product_row.quantities[size_index]
        if isinstance(existing, (int, float)) and isinstance(quantity, (int, float)):
            product_row.quantities[size_index] = existing + quantity
        else:
            product_row.quantities[size_index] = quantity
        if isinstance(product_row.total, (int, float)) and isinstance(quantity, (int, float)):
            product_row.total += quantity

        image_bytes = images_by_row.get(row_index)
        if image_bytes:
            images_by_style.setdefault(style_no, image_bytes)

    for product_row in grouped.values():
        while len(product_row.quantities) < len(sizes):
            product_row.quantities.append(None)

    return sizes, list(grouped.values()), images_by_style


def extract_pdf_text_items(page) -> list[tuple[float, float, str]]:
    items: list[tuple[float, float, str]] = []

    def visitor(text, _cm, tm, _font_dict, _font_size):
        cleaned = clean_text(text)
        if cleaned:
            items.append((float(tm[4]), float(tm[5]), cleaned))

    page.extract_text(visitor_text=visitor)
    return items


def extract_pdf_image_placements(reader: PdfReader, page) -> list[tuple[float, bytes]]:
    images = {f"/{image.name.rsplit('.', 1)[0]}": image.data for image in page.images}
    placements: list[tuple[float, bytes]] = []
    last_cm: list[Any] | None = None
    content = page.get_contents()
    if content is None:
        return placements

    stream = ContentStream(content, reader)
    for operands, operator in stream.operations:
        if operator == b"cm":
            last_cm = operands
        elif operator == b"Do" and last_cm:
            image_name = str(operands[0])
            image_bytes = images.get(image_name)
            if image_bytes:
                placements.append((float(last_cm[5]), image_bytes))
    return placements


def join_fragments(items: list[tuple[float, float, str]]) -> str:
    return "".join(text for _x, _y, text in sorted(items, key=lambda item: (-item[1], item[0]))).strip()


def row_bands(row_anchors: list[tuple[float, float, str]]) -> list[tuple[float, float, float, str]]:
    anchors = sorted(row_anchors, key=lambda item: -item[1])
    bands: list[tuple[float, float, float, str]] = []
    for index, (_x, y, seq) in enumerate(anchors):
        upper = y + 18 if index == 0 else (anchors[index - 1][1] + y) / 2
        lower = y - 30 if index == len(anchors) - 1 else (y + anchors[index + 1][1]) / 2
        bands.append((upper, lower, y, seq))
    return bands


def parse_pdf_row(items: list[tuple[float, float, str]], anchor_y: float) -> ProductRow | None:
    seq = join_fragments([(x, y, text) for x, y, text in items if 35 <= x < 65])
    order_no = join_fragments([(x, y, text) for x, y, text in items if 65 <= x < 105])
    customer = join_fragments([(x, y, text) for x, y, text in items if 105 <= x < 155])
    style = join_fragments([(x, y, text) for x, y, text in items if 188 <= x < 236])
    name = join_fragments([(x, y, text) for x, y, text in items if 236 <= x < 286])
    color = join_fragments([(x, y, text) for x, y, text in items if 286 <= x < 336])
    total_text = join_fragments([(x, y, text) for x, y, text in items if x >= 745 and re.fullmatch(r"-?\d+(\.\d+)?", text)])
    if not style:
        return None

    quantities: list[int | float | str | None] = [None] * len(SIZE_HEADERS)
    for x, _y, text in items:
        if not (336 <= x < 745) or not re.fullmatch(r"-?\d+(\.\d+)?", text):
            continue
        nearest = min(range(len(SIZE_X_POSITIONS)), key=lambda idx: abs(x - SIZE_X_POSITIONS[idx]))
        quantities[nearest] = parse_number(text)

    return ProductRow(
        row_index=int(round(anchor_y * 100)),
        seq=seq,
        order_no=order_no,
        customer=customer,
        style_no=style,
        name=name,
        color=color,
        quantities=quantities,
        total=parse_number(total_text),
    )


def load_pdf_groups(input_path: Path) -> tuple[list[str], OrderedDict[str, ProductGroup]]:
    reader = PdfReader(str(input_path))
    groups: OrderedDict[str, ProductGroup] = OrderedDict()

    for page in reader.pages:
        text_items = extract_pdf_text_items(page)
        anchors = [
            (x, y, text)
            for x, y, text in text_items
            if 35 <= x <= 55 and 50 < y < 485 and text.isdigit()
        ]
        placements = extract_pdf_image_placements(reader, page)

        for upper, lower, anchor_y, _seq in row_bands(anchors):
            band_items = [(x, y, text) for x, y, text in text_items if lower <= y <= upper]
            product_row = parse_pdf_row(band_items, anchor_y)
            if not product_row:
                continue
            if product_row.style_no == "00000" or product_row.name == "抹零":
                continue

            group = groups.setdefault(
                product_row.style_no,
                ProductGroup(style_no=product_row.style_no, name=product_row.name),
            )
            if not group.name and product_row.name:
                group.name = product_row.name
            if group.image_bytes is None:
                for image_y, image_bytes in placements:
                    if lower <= image_y <= upper:
                        group.image_bytes = image_bytes
                        break
            group.rows.append(product_row)

    return SIZE_HEADERS, groups


def load_pdf_rows(input_path: Path) -> tuple[list[str], list[ProductRow], dict[str, bytes]]:
    reader = PdfReader(str(input_path))
    rows: list[ProductRow] = []
    images_by_style: dict[str, bytes] = {}

    for page in reader.pages:
        text_items = extract_pdf_text_items(page)
        anchors = [
            (x, y, text)
            for x, y, text in text_items
            if 35 <= x <= 55 and 50 < y < 485 and text.isdigit()
        ]
        placements = extract_pdf_image_placements(reader, page)

        for upper, lower, anchor_y, _seq in row_bands(anchors):
            band_items = [(x, y, text) for x, y, text in text_items if lower <= y <= upper]
            product_row = parse_pdf_row(band_items, anchor_y)
            if not product_row:
                continue
            if product_row.style_no == "00000" or product_row.name == "抹零":
                continue
            for image_y, image_bytes in placements:
                if lower <= image_y <= upper:
                    images_by_style.setdefault(product_row.style_no, image_bytes)
                    break
            rows.append(product_row)

    rows.sort(key=lambda item: int(item.seq) if item.seq.isdigit() else item.row_index)
    return SIZE_HEADERS, rows, images_by_style


def style_cell(cell, fill=None, bold=False, color="000000", center=True):
    cell.border = BORDER
    cell.font = Font(name="Microsoft YaHei", size=10, bold=bold, color=color)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)


def add_group_image(ws, row: int, image_bytes: bytes | None, group_height: int, column: int = 1):
    if not image_bytes:
        return
    try:
        image = ExcelImage(io.BytesIO(image_bytes))
        image_copy = copy.copy(image)
        max_width = 120
        cell_area_height = max(DATA_ROW_HEIGHT_PX, group_height * DATA_ROW_HEIGHT_PX)
        max_height = max(24, cell_area_height - 8)
        ratio = min(max_width / image_copy.width, max_height / image_copy.height, 1)
        width = int(image_copy.width * ratio)
        height = int(image_copy.height * ratio)
        image_copy.width = width
        image_copy.height = height
        row_offset = max(4, int((cell_area_height - height) / 2))
        start = AnchorMarker(
            col=column - 1,
            row=row - 1,
            colOff=pixels_to_EMU(4),
            rowOff=pixels_to_EMU(row_offset),
        )
        image_copy.anchor = OneCellAnchor(
            _from=start,
            ext=XDRPositiveSize2D(cx=pixels_to_EMU(width), cy=pixels_to_EMU(height)),
        )
        ws.add_image(image_copy)
    except Exception:
        return


def merge_range_if_needed(ws, start_row: int, end_row: int, col: int) -> None:
    if end_row <= start_row:
        return
    ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
    ws.cell(start_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def contiguous_runs(rows: list[ProductRow], key_name: str) -> list[tuple[int, int, str]]:
    runs: list[tuple[int, int, str]] = []
    if not rows:
        return runs
    start = 0
    previous = clean_text(getattr(rows[0], key_name))
    for index, product_row in enumerate(rows[1:], start=1):
        current = clean_text(getattr(product_row, key_name))
        if current != previous:
            runs.append((start, index - 1, previous))
            start = index
            previous = current
    runs.append((start, len(rows) - 1, previous))
    return runs


def write_output(size_headers: list[str], groups: OrderedDict[str, ProductGroup], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "整理结果"

    fixed_headers = ["图片", "款号", "名称", "序号", "订单号", "客户", "颜色"]
    total_col = len(fixed_headers) + len(size_headers) + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)
    title = ws.cell(1, 1, "销售订货明细整理表")
    title.fill = TITLE_FILL
    title.font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(2, col, header)
        style_cell(cell, GROUP_FILL, bold=True)
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

    size_start = len(fixed_headers) + 1
    size_end = size_start + len(size_headers) - 1
    ws.merge_cells(start_row=2, start_column=size_start, end_row=2, end_column=size_end)
    size_title = ws.cell(2, size_start, "尺码 / 数量")
    style_cell(size_title, GROUP_FILL, bold=True)

    for offset, header in enumerate(size_headers):
        cell = ws.cell(3, size_start + offset, header)
        style_cell(cell, SIZE_FILL, bold=True)

    total_header = ws.cell(2, total_col, "订货数")
    style_cell(total_header, GROUP_FILL, bold=True)
    ws.merge_cells(start_row=2, start_column=total_col, end_row=3, end_column=total_col)

    row = 4
    for group in groups.values():
        start_row = row
        for product_row in group.rows:
            quantities = [hide_zero(value) for value in product_row.quantities]
            values = [
                "",
                group.style_no,
                group.name,
                product_row.seq,
                product_row.order_no,
                product_row.customer,
                product_row.color,
                *quantities,
                product_row.total,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row, col, value)
                style_cell(cell, center=(col not in (3, 7)))
            ws.row_dimensions[row].height = DATA_ROW_HEIGHT_PT
            row += 1

        end_row = row - 1
        if end_row > start_row:
            for col in (1, 2, 3):
                merge_range_if_needed(ws, start_row, end_row, col)

        add_group_image(ws, start_row, group.image_bytes, len(group.rows))

    widths = {
        "A": 18,
        "B": 14,
        "C": 28,
        "D": 8,
        "E": 12,
        "F": 14,
        "G": 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    for col in range(size_start, total_col):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions[get_column_letter(total_col)].width = 12

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(total_col)}{max(row - 1, 3)}"
    ws.sheet_view.showGridLines = False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_output_rows(
    size_headers: list[str],
    rows: list[ProductRow],
    images_by_style: dict[str, bytes],
    output_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "整理结果"

    fixed_headers = ["订单号", "客户", "图片", "款号", "名称", "颜色"]
    total_col = len(fixed_headers) + len(size_headers) + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)
    title = ws.cell(1, 1, "销售订货明细整理表")
    title.fill = TITLE_FILL
    title.font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(2, col, header)
        style_cell(cell, GROUP_FILL, bold=True)
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

    size_start = len(fixed_headers) + 1
    size_end = size_start + len(size_headers) - 1
    ws.merge_cells(start_row=2, start_column=size_start, end_row=2, end_column=size_end)
    style_cell(ws.cell(2, size_start, "尺码 / 数量"), GROUP_FILL, bold=True)

    for offset, header in enumerate(size_headers):
        style_cell(ws.cell(3, size_start + offset, header), SIZE_FILL, bold=True)

    total_header = ws.cell(2, total_col, "订货数")
    style_cell(total_header, GROUP_FILL, bold=True)
    ws.merge_cells(start_row=2, start_column=total_col, end_row=3, end_column=total_col)

    row_index = 4
    for product_row in rows:
        quantities = [hide_zero(value) for value in product_row.quantities]
        values = [
            product_row.order_no,
            product_row.customer,
            "",
            product_row.style_no,
            product_row.name,
            product_row.color,
            *quantities,
            product_row.total,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col, value)
            style_cell(cell, center=(col not in (5, 6)))
        ws.row_dimensions[row_index].height = DATA_ROW_HEIGHT_PT
        row_index += 1

    first_data_row = 4
    used_images: set[str] = set()
    for start, end, style_no in contiguous_runs(rows, "style_no"):
        start_row = first_data_row + start
        end_row = first_data_row + end
        merge_range_if_needed(ws, start_row, end_row, 3)
        merge_range_if_needed(ws, start_row, end_row, 4)
        merge_range_if_needed(ws, start_row, end_row, 5)
        if style_no and style_no not in used_images:
            add_group_image(ws, start_row, images_by_style.get(style_no), end - start + 1, column=3)
            used_images.add(style_no)

    for key_name, col in (("order_no", 1), ("customer", 2)):
        for start, end, value in contiguous_runs(rows, key_name):
            if value:
                merge_range_if_needed(ws, first_data_row + start, first_data_row + end, col)

    for col_letter, width in {
        "A": 12,
        "B": 14,
        "C": 18,
        "D": 14,
        "E": 28,
        "F": 16,
    }.items():
        ws.column_dimensions[col_letter].width = width
    for col in range(size_start, total_col):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions[get_column_letter(total_col)].width = 12

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(total_col)}{max(row_index - 1, 3)}"
    ws.sheet_view.showGridLines = False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def format_excel(input_path: Path, output_path: Path | None = None) -> Path:
    input_path = input_path.expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"文件不存在：{input_path}")
    if output_path is None:
        output_path = input_path.with_name(f"{input_path.stem}_整理后.xlsx")
    else:
        output_path = output_path.expanduser().resolve()

    if input_path.suffix.lower() == ".xlsx":
        workbook = openpyxl.load_workbook(input_path, read_only=True)
        ws = workbook.active
        header_names = {clean_text(cell.value) for cell in ws[1] if clean_text(cell.value)}
        workbook.close()
        if {"尺码", "订货数", "款号", "颜色"}.issubset(header_names):
            size_headers, rows, images_by_style = load_vertical_xlsx_rows(input_path)
            if not rows:
                raise ValueError("没有找到可整理的商品行。")
            write_output_rows(size_headers, rows, images_by_style, output_path)
        else:
            size_headers, groups = load_groups(input_path)
            if not groups:
                raise ValueError("没有找到可整理的商品行。")
            write_output(size_headers, groups, output_path)
    elif input_path.suffix.lower() == ".pdf":
        size_headers, rows, images_by_style = load_pdf_rows(input_path)
        if not rows:
            raise ValueError("没有找到可整理的商品行。")
        write_output_rows(size_headers, rows, images_by_style, output_path)
    else:
        raise ValueError("当前版本支持 .xlsx 和同版式 PDF 文件。")
    return output_path


def run_gui() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.title("Excel整理")
    root.geometry("560x270")
    root.resizable(False, False)

    selected = tk.StringVar()
    status = tk.StringVar(value="请选择需要整理的 Excel 文件。")

    def choose_file():
        filename = filedialog.askopenfilename(
            title="选择投票订单文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("PDF 文件", "*.pdf"), ("所有文件", "*.*")],
        )
        if filename:
            selected.set(filename)
            status.set("已选择文件，点击“开始整理”。")

    def start():
        if not selected.get():
            messagebox.showwarning("提示", "请先选择一个Excel或PDF文件。")
            return
        try:
            status.set("正在整理，请稍等...")
            root.update_idletasks()
            output = format_excel(Path(selected.get()))
            status.set(f"整理完成：{output.name}")
            messagebox.showinfo("完成", f"整理完成：\n{output}")
        except Exception as exc:
            status.set("整理失败，请检查文件格式。")
            messagebox.showerror("整理失败", str(exc))

    tk.Label(root, text="Excel整理", font=("Microsoft YaHei", 18, "bold")).pack(pady=(20, 8))
    tk.Label(root, text="按相同款号合并图片，重新排版为“尺码 / 数量”格式。", font=("Microsoft YaHei", 10)).pack()
    tk.Label(root, text="当前支持 .xlsx 和同版式销售订货明细 PDF。", font=("Microsoft YaHei", 9), fg="#666666").pack(pady=(3, 0))
    tk.Entry(root, textvariable=selected, width=72).pack(pady=18, padx=18)

    button_frame = tk.Frame(root)
    button_frame.pack()
    tk.Button(button_frame, text="选择文件", width=14, command=choose_file).pack(side="left", padx=8)
    tk.Button(button_frame, text="开始整理", width=14, command=start).pack(side="left", padx=8)
    tk.Label(root, textvariable=status, font=("Microsoft YaHei", 9), fg="#1F4E78").pack(pady=(18, 0))

    root.mainloop()


def main() -> int:
    parser = argparse.ArgumentParser(description="按款号整理投票Excel，重复款号只保留一张图片。")
    parser.add_argument("input", nargs="?", help="输入Excel文件路径")
    parser.add_argument("-o", "--output", help="输出Excel文件路径")
    parser.add_argument("--gui", action="store_true", help="打开图形界面")
    args = parser.parse_args()

    if args.gui or not args.input:
        run_gui()
        return 0

    output = format_excel(Path(args.input), Path(args.output) if args.output else None)
    print(f"整理完成：{output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
